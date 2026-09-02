from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from uuid import uuid4

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import connections, transaction
from django.shortcuts import redirect, render
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook

from control.gf_authz.permissions import gf_has_perm
from .services.entity_access import require_tenant_context


HEADER_ALIASES = {
    "written_date": {"작성일", "작성일자", "거래일자"},
    "issued_date": {"발행일", "발행일자", "발급일", "발급일자"},
    "invoice_type": {"구분", "매출매입구분", "계산서구분", "세금계산서구분"},
    "partner": {"거래처", "거래처명", "업체명", "상호"},
    "supplier_name": {"공급자상호", "공급자상호명", "공급자명", "공급자"},
    "supplier_biz_no": {"공급자사업자등록번호", "공급자사업자번호", "공급자등록번호"},
    "recipient_name": {"공급받는자상호", "공급받는자상호명", "공급받는자명", "공급받는자"},
    "recipient_biz_no": {"공급받는자사업자등록번호", "공급받는자사업자번호", "공급받는자등록번호"},
    "contract": {"계약", "계약명", "계약번호"},
    "supply_amount": {"공급가액", "공급가", "공급가액합계"},
    "vat_amount": {"부가세", "세액", "세액합계", "부가가치세", "vat"},
    "total_amount": {"합계", "합계금액", "총액", "총합계"},
    "approval_no": {"계산서번호", "승인번호", "계산서번호승인번호", "전자세금계산서승인번호"},
    "status": {"상태"},
    "transaction_date": {"거래일자", "거래일", "일자", "거래일시"},
    "transaction_type": {"입출금구분", "입금출금구분", "구분", "거래구분"},
    "amount": {"금액", "거래금액", "입출금액"},
    "account": {"계좌", "계좌명", "계좌번호"},
    "description": {"적요", "내용", "거래내용"},
    "category": {"분류", "수입지출분류", "계정분류"},
    "evidence_type": {"증빙", "증빙유형"},
    "memo": {"비고", "메모"},
}

INVOICE_FIELDS = {
    "written_date", "issued_date", "invoice_type", "partner", "supplier_name", "supplier_biz_no",
    "recipient_name", "recipient_biz_no", "contract", "supply_amount", "vat_amount", "total_amount",
    "approval_no", "status", "memo",
}
TRANSACTION_FIELDS = {
    "transaction_date", "transaction_type", "amount", "partner", "contract", "account", "description",
    "category", "evidence_type", "memo",
}


def _norm(value):
    text = str(value or "").strip().lower().replace("㈜", "주식회사")
    return re.sub(r"[\s·/()_\-.:]+", "", text)


def _norm_biz(value):
    return re.sub(r"\D", "", str(value or ""))


def _norm_company(value):
    text = str(value or "").strip()
    text = re.sub(r"^\s*(?:\(\s*[주유합]\s*\)|㈜|주식회사|유한회사|합자회사|합명회사)\s*", "", text)
    text = re.sub(r"\s*(?:\(\s*[주유합]\s*\)|㈜|주식회사|유한회사|합자회사|합명회사)\s*$", "", text)
    return re.sub(r"\s+", "", text).lower()


def _headers(row):
    normalized = {_norm(v): idx for idx, v in enumerate(row) if v not in (None, "")}
    result = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if _norm(alias) in normalized:
                result[field] = normalized[_norm(alias)]
                break
    return result


def _cell(row, mapping, key):
    idx = mapping.get(key)
    return row[idx] if idx is not None and idx < len(row) else None


def _date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace(".", "-").replace("/", "-")
    if " " in text:
        text = text.split(" ", 1)[0]
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%y-%m-%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"날짜 형식을 확인하세요: {value}")


def _money(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "").replace("₩", "").strip()).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        raise ValueError(f"금액 형식을 확인하세요: {value}")


def _system_code(alias, system_key):
    with connections[alias].cursor() as cur:
        cur.execute("SELECT code FROM ops.settings_nodes WHERE system_key=%s AND active=true LIMIT 1", [system_key])
        row = cur.fetchone()
    return row[0] if row else None


def _ref_code(alias, field_ref, value):
    text = str(value or "").strip()
    if not text:
        return None
    with connections[alias].cursor() as cur:
        cur.execute(
            """
            SELECT v.code
              FROM ops.settings_nodes c
              JOIN ops.settings_nodes v ON v.parent_id=c.id
             WHERE c.field_ref=%s AND v.active=true
               AND (lower(v.code)=lower(%s) OR lower(v.name)=lower(%s))
             ORDER BY CASE WHEN lower(v.code)=lower(%s) THEN 0 ELSE 1 END, v.ord
             LIMIT 1
            """,
            [field_ref, text, text, text],
        )
        row = cur.fetchone()
    return row[0] if row else None


def _actor(request):
    user = request.user
    return str(getattr(user, "email", None) or getattr(user, "username", None) or getattr(user, "pk", ""))[:255]


def _default_project_id(alias, contract_id):
    if not contract_id:
        return None
    with connections[alias].cursor() as cur:
        cur.execute("SELECT id::text FROM prj.projects WHERE contract_id=%s ORDER BY id LIMIT 2", [contract_id])
        rows = cur.fetchall()
    return rows[0][0] if len(rows) == 1 else None


def _contract_id(alias, value):
    text = str(value or "").strip()
    if not text:
        return None
    with connections[alias].cursor() as cur:
        cur.execute(
            "SELECT id::text FROM ctr.contracts WHERE lower(COALESCE(code,''))=lower(%s) OR lower(COALESCE(name,''))=lower(%s) LIMIT 2",
            [text, text],
        )
        rows = cur.fetchall()
    return rows[0][0] if len(rows) == 1 else None


def _account_id(alias, value):
    text = str(value or "").strip()
    if not text:
        return None
    with connections[alias].cursor() as cur:
        cur.execute(
            """
            SELECT id::text FROM fin.accounts
             WHERE active=true AND (lower(account_name)=lower(%s) OR replace(COALESCE(account_number,''),'-','')=replace(%s,'-',''))
             LIMIT 2
            """,
            [text, text],
        )
        rows = cur.fetchall()
    return rows[0][0] if len(rows) == 1 else None


def _partner_indexes(alias):
    with connections[alias].cursor() as cur:
        cur.execute("SELECT id::text,name,COALESCE(biz_no,'') FROM ctr.partners WHERE COALESCE(status,'') NOT IN ('inactive','deleted')")
        rows = cur.fetchall()
    by_biz = {}
    by_name = {}
    for partner_id, name, biz_no in rows:
        biz = _norm_biz(biz_no)
        norm_name = _norm_company(name)
        if biz:
            by_biz.setdefault(biz, []).append((partner_id, name))
        if norm_name:
            by_name.setdefault(norm_name, []).append((partner_id, name))
    return by_biz, by_name


def _partner_match(by_biz, by_name, name, biz_no):
    biz = _norm_biz(biz_no)
    if biz and len(by_biz.get(biz, [])) == 1:
        return by_biz[biz][0][0]
    norm_name = _norm_company(name)
    if norm_name and len(by_name.get(norm_name, [])) == 1:
        return by_name[norm_name][0][0]
    return None


def _own_biz_numbers(alias):
    try:
        with connections[alias].cursor() as cur:
            cur.execute("SELECT COALESCE(biz_no,'') FROM ops.my_org_units WHERE COALESCE(biz_no,'')<>''")
            return {_norm_biz(r[0]) for r in cur.fetchall() if _norm_biz(r[0])}
    except Exception:
        return set()


def _find_header(ws, import_type, manual_row=None):
    relevant = INVOICE_FIELDS if import_type == "invoice" else TRANSACTION_FIELDS
    if manual_row:
        row = next(ws.iter_rows(min_row=manual_row, max_row=manual_row, values_only=True), None) or []
        mapping = _headers(row)
        return manual_row, mapping, list(row)

    best = None
    max_row = min(int(ws.max_row or 1), 20)
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        mapping = _headers(row)
        score = len(set(mapping).intersection(relevant))
        if best is None or score > best[0]:
            best = (score, row_no, mapping, list(row))
    if not best or best[0] < 2:
        raise ValueError("상단 20행에서 헤더를 자동으로 찾지 못했습니다. 헤더 행 번호를 직접 입력해 주세요.")
    return best[1], best[2], best[3]


def _hash(parts):
    value = "|".join(str(v or "").strip().lower() for v in parts)
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _existing_invoice(alias, payload):
    with connections[alias].cursor() as cur:
        if payload.get("approval_no"):
            cur.execute(
                """
                SELECT i.id::text,COALESCE(i.issued_date,i.written_date),COALESCE(p.name,i.source_partner_name,''),COALESCE(c.name,''),i.total_amount
                  FROM fin.tax_invoices i
                  LEFT JOIN ctr.partners p ON p.id=i.partner_id
                  LEFT JOIN ctr.contracts c ON c.id=i.contract_id
                 WHERE i.is_deleted=false AND i.approval_no=%s
                 ORDER BY i.created_at DESC LIMIT 1
                """,
                [payload["approval_no"]],
            )
        elif payload.get("partner_id"):
            cur.execute(
                """
                SELECT i.id::text,COALESCE(i.issued_date,i.written_date),COALESCE(p.name,i.source_partner_name,''),COALESCE(c.name,''),i.total_amount
                  FROM fin.tax_invoices i
                  LEFT JOIN ctr.partners p ON p.id=i.partner_id
                  LEFT JOIN ctr.contracts c ON c.id=i.contract_id
                 WHERE i.is_deleted=false AND COALESCE(i.issued_date,i.written_date) IS NOT DISTINCT FROM %s::date
                   AND i.partner_id=%s AND i.supply_amount=%s AND i.vat_amount=%s AND i.total_amount=%s
                 ORDER BY i.created_at DESC LIMIT 1
                """,
                [payload.get("display_date") or None, payload["partner_id"], payload["supply_amount"], payload["vat_amount"], payload["total_amount"]],
            )
        else:
            cur.execute(
                """
                SELECT i.id::text,COALESCE(i.issued_date,i.written_date),COALESCE(p.name,i.source_partner_name,''),COALESCE(c.name,''),i.total_amount
                  FROM fin.tax_invoices i
                  LEFT JOIN ctr.partners p ON p.id=i.partner_id
                  LEFT JOIN ctr.contracts c ON c.id=i.contract_id
                 WHERE i.is_deleted=false AND COALESCE(i.issued_date,i.written_date) IS NOT DISTINCT FROM %s::date
                   AND lower(COALESCE(i.source_partner_name,''))=lower(%s)
                   AND i.supply_amount=%s AND i.vat_amount=%s AND i.total_amount=%s
                 ORDER BY i.created_at DESC LIMIT 1
                """,
                [payload.get("display_date") or None, payload.get("source_partner_name") or "", payload["supply_amount"], payload["vat_amount"], payload["total_amount"]],
            )
        row = cur.fetchone()
    if not row:
        return None
    return {"id": row[0], "date": row[1].isoformat() if row[1] else "", "partner": row[2], "contract": row[3], "amount": str(row[4] or 0)}


def _existing_transaction(alias, payload):
    with connections[alias].cursor() as cur:
        cur.execute(
            """
            SELECT t.id::text,t.transaction_date,COALESCE(p.name,t.source_partner_name,''),COALESCE(c.name,''),t.amount,COALESCE(t.description,'')
              FROM fin.transactions t
              LEFT JOIN ctr.partners p ON p.id=t.partner_id
              LEFT JOIN ctr.contracts c ON c.id=t.contract_id
             WHERE t.is_deleted=false AND t.transaction_date=%s::date AND t.transaction_type=%s AND t.amount=%s
               AND lower(COALESCE(t.description,''))=lower(%s)
             ORDER BY t.created_at DESC LIMIT 1
            """,
            [payload["transaction_date"], payload["transaction_type"], payload["amount"], payload.get("description") or ""],
        )
        row = cur.fetchone()
    if not row:
        return None
    return {"id": row[0], "date": row[1].isoformat() if row[1] else "", "partner": row[2], "contract": row[3], "amount": str(row[4] or 0), "description": row[5]}


def _invoice_preview(alias, ws, mapping, header_row, direction_pref):
    by_biz, by_name = _partner_indexes(alias)
    own_biz = _own_biz_numbers(alias)
    sales_code = _system_code(alias, "finance.value.invoice_type.sales")
    purchase_code = _system_code(alias, "finance.value.invoice_type.purchase")
    issued_code = _system_code(alias, "finance.value.invoice_status.issued")
    rows = []

    for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(v not in (None, "") for v in row):
            continue
        if len(rows) >= 1000:
            break
        warnings = []
        errors = []
        try:
            written = _date(_cell(row, mapping, "written_date"))
            issued = _date(_cell(row, mapping, "issued_date"))
        except ValueError as exc:
            written = issued = None
            warnings.append(str(exc))

        supplier_name = str(_cell(row, mapping, "supplier_name") or "").strip()
        supplier_biz = _norm_biz(_cell(row, mapping, "supplier_biz_no"))
        recipient_name = str(_cell(row, mapping, "recipient_name") or "").strip()
        recipient_biz = _norm_biz(_cell(row, mapping, "recipient_biz_no"))
        generic_partner = str(_cell(row, mapping, "partner") or "").strip()

        invoice_type = _ref_code(alias, "finance.invoice_type", _cell(row, mapping, "invoice_type"))
        if not invoice_type:
            if supplier_biz and supplier_biz in own_biz:
                invoice_type = sales_code
            elif recipient_biz and recipient_biz in own_biz:
                invoice_type = purchase_code
            elif direction_pref == "sales":
                invoice_type = sales_code
            elif direction_pref == "purchase":
                invoice_type = purchase_code

        if invoice_type == sales_code:
            source_partner_name, source_partner_biz = recipient_name or generic_partner, recipient_biz
        elif invoice_type == purchase_code:
            source_partner_name, source_partner_biz = supplier_name or generic_partner, supplier_biz
        else:
            source_partner_name, source_partner_biz = generic_partner or recipient_name or supplier_name, recipient_biz or supplier_biz
            errors.append("매출/매입 구분을 판정할 수 없습니다. 업로드 옵션에서 방향을 선택해 다시 미리보기 하세요.")

        partner_id = _partner_match(by_biz, by_name, source_partner_name, source_partner_biz)
        if source_partner_name and not partner_id:
            warnings.append("GeoFlow 거래처와 자동 매칭되지 않았습니다. 거래처 미연결 상태로 저장할 수 있습니다.")

        contract_id = _contract_id(alias, _cell(row, mapping, "contract"))
        project_id = _default_project_id(alias, contract_id)

        try:
            supply = _money(_cell(row, mapping, "supply_amount"))
            vat = _money(_cell(row, mapping, "vat_amount"))
            total = _money(_cell(row, mapping, "total_amount"))
            if supply is None and total is not None:
                supply = (total / Decimal("1.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                vat = total - supply
            elif supply is not None and vat is None:
                vat = (supply * Decimal("0.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            if total is None and supply is not None:
                total = supply + (vat or Decimal("0"))
            if supply is None and total is None:
                errors.append("공급가액 또는 합계금액이 없습니다.")
                supply = vat = total = Decimal("0")
            if vat is None:
                vat = Decimal("0")
        except ValueError as exc:
            errors.append(str(exc))
            supply = vat = total = Decimal("0")

        approval_no = str(_cell(row, mapping, "approval_no") or "").strip() or None
        status = _ref_code(alias, "finance.invoice_status", _cell(row, mapping, "status")) or issued_code
        memo = str(_cell(row, mapping, "memo") or "").strip()
        display_date = issued or written
        payload = {
            "written_date": written.isoformat() if written else "",
            "issued_date": issued.isoformat() if issued else "",
            "display_date": display_date.isoformat() if display_date else "",
            "invoice_type": invoice_type or "",
            "partner_id": partner_id or "",
            "source_partner_name": source_partner_name,
            "source_partner_biz_no": source_partner_biz,
            "contract_id": contract_id or "",
            "project_id": project_id or "",
            "supply_amount": str(supply), "vat_amount": str(vat), "total_amount": str(total),
            "approval_no": approval_no or "", "status": status or "", "memo": memo,
        }
        payload["fingerprint"] = _hash([invoice_type, payload["display_date"], source_partner_biz or _norm_company(source_partner_name), supply, vat, total])
        duplicate = _existing_invoice(alias, payload) if not errors else None
        status_key = "error" if errors else ("duplicate" if duplicate else ("warning" if warnings else "new"))
        rows.append({
            "index": len(rows), "row_no": row_no, "status": status_key,
            "status_label": {"new": "신규", "warning": "확인", "duplicate": "중복 의심", "error": "저장 불가"}[status_key],
            "message": " / ".join(errors or warnings), "duplicate": duplicate, "payload": payload,
            "display": {
                "date": payload["display_date"], "partner": source_partner_name or "-", "contract": str(_cell(row, mapping, "contract") or "-"),
                "supply": str(supply), "vat": str(vat), "total": str(total), "approval_no": approval_no or "-",
            },
            "selectable": not errors, "default_selected": not errors and not duplicate,
        })
    return rows


def _transaction_preview(alias, ws, mapping, header_row):
    by_biz, by_name = _partner_indexes(alias)
    rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(v not in (None, "") for v in row):
            continue
        if len(rows) >= 1000:
            break
        warnings = []
        errors = []
        try:
            tx_date = _date(_cell(row, mapping, "transaction_date"))
        except ValueError as exc:
            tx_date = None
            errors.append(str(exc))
        if not tx_date:
            errors.append("거래일자가 없습니다.")

        tx_type = _ref_code(alias, "finance.transaction_type", _cell(row, mapping, "transaction_type"))
        if not tx_type:
            errors.append("입금/출금 구분을 찾을 수 없습니다.")
        try:
            amount = _money(_cell(row, mapping, "amount"))
        except ValueError as exc:
            amount = None
            errors.append(str(exc))
        if amount is None:
            errors.append("금액이 없습니다.")
            amount = Decimal("0")

        partner_name = str(_cell(row, mapping, "partner") or "").strip()
        partner_id = _partner_match(by_biz, by_name, partner_name, "")
        if partner_name and not partner_id:
            warnings.append("GeoFlow 거래처와 자동 매칭되지 않았습니다.")

        contract_id = _contract_id(alias, _cell(row, mapping, "contract"))
        project_id = _default_project_id(alias, contract_id)
        account_id = _account_id(alias, _cell(row, mapping, "account"))
        category = _ref_code(alias, "finance.transaction_category", _cell(row, mapping, "category"))
        evidence = _ref_code(alias, "finance.evidence_type", _cell(row, mapping, "evidence_type"))
        description = str(_cell(row, mapping, "description") or "").strip()
        memo = str(_cell(row, mapping, "memo") or "").strip()
        payload = {
            "transaction_date": tx_date.isoformat() if tx_date else "",
            "transaction_type": tx_type or "", "amount": str(amount), "partner_id": partner_id or "",
            "source_partner_name": partner_name, "source_partner_biz_no": "", "account_id": account_id or "",
            "description": description, "contract_id": contract_id or "", "project_id": project_id or "",
            "category_code": category or "", "evidence_type": evidence or "", "memo": memo,
        }
        payload["fingerprint"] = _hash([payload["transaction_date"], tx_type, amount, _norm_company(partner_name), description])
        duplicate = _existing_transaction(alias, payload) if not errors else None
        status_key = "error" if errors else ("duplicate" if duplicate else ("warning" if warnings else "new"))
        rows.append({
            "index": len(rows), "row_no": row_no, "status": status_key,
            "status_label": {"new": "신규", "warning": "확인", "duplicate": "중복 의심", "error": "저장 불가"}[status_key],
            "message": " / ".join(errors or warnings), "duplicate": duplicate, "payload": payload,
            "display": {"date": payload["transaction_date"], "partner": partner_name or "-", "contract": str(_cell(row, mapping, "contract") or "-"), "description": description or "-", "amount": str(amount)},
            "selectable": not errors, "default_selected": not errors and not duplicate,
        })
    return rows


def _commit_preview(alias, preview, selected, request):
    batch_id = str(uuid4())
    created = 0
    skipped = 0
    with transaction.atomic(using=alias):
        with connections[alias].cursor() as cur:
            for row in preview.get("rows", []):
                if str(row.get("index")) not in selected or not row.get("selectable"):
                    continue
                p = row.get("payload") or {}
                if preview.get("import_type") == "invoice":
                    cur.execute(
                        """
                        INSERT INTO fin.tax_invoices(
                            written_date,issued_date,invoice_type,partner_id,contract_id,project_id,
                            supply_amount,vat_amount,total_amount,approval_no,status,memo,created_by,
                            source_type,source_partner_name,source_partner_biz_no,import_fingerprint,import_batch_id
                        ) VALUES(
                            NULLIF(%s,'')::date,NULLIF(%s,'')::date,%s,NULLIF(%s,'')::uuid,NULLIF(%s,'')::uuid,NULLIF(%s,'')::uuid,
                            %s,%s,%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s,
                            'xlsx',NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s::uuid
                        )
                        """,
                        [
                            p.get("written_date", ""), p.get("issued_date", ""), p.get("invoice_type", ""), p.get("partner_id", ""),
                            p.get("contract_id", ""), p.get("project_id", ""), p.get("supply_amount", "0"), p.get("vat_amount", "0"), p.get("total_amount", "0"),
                            p.get("approval_no", ""), p.get("status", ""), p.get("memo", ""), _actor(request), p.get("source_partner_name", ""),
                            p.get("source_partner_biz_no", ""), p.get("fingerprint", ""), batch_id,
                        ],
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO fin.transactions(
                            transaction_date,transaction_type,amount,partner_id,account_id,description,contract_id,project_id,
                            category_code,evidence_type,memo,created_by,source_type,source_partner_name,source_partner_biz_no,import_fingerprint,import_batch_id
                        ) VALUES(
                            %s::date,%s,%s,NULLIF(%s,'')::uuid,NULLIF(%s,'')::uuid,NULLIF(%s,''),NULLIF(%s,'')::uuid,NULLIF(%s,'')::uuid,
                            NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s,'xlsx',NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s::uuid
                        )
                        """,
                        [
                            p.get("transaction_date", ""), p.get("transaction_type", ""), p.get("amount", "0"), p.get("partner_id", ""), p.get("account_id", ""),
                            p.get("description", ""), p.get("contract_id", ""), p.get("project_id", ""), p.get("category_code", ""), p.get("evidence_type", ""),
                            p.get("memo", ""), _actor(request), p.get("source_partner_name", ""), p.get("source_partner_biz_no", ""), p.get("fingerprint", ""), batch_id,
                        ],
                    )
                created += 1
    skipped = len([r for r in preview.get("rows", []) if r.get("selectable")]) - created
    return created, skipped


def _render(request, **context):
    context.setdefault("preview", None)
    context.setdefault("import_type", "invoice")
    context.setdefault("invoice_direction", "auto")
    context.setdefault("header_row", "")
    return render(request, "geoflow_ops/finance/finance_import.html", context)


@login_required
@require_http_methods(["GET", "POST"])
def finance_import(request):
    alias = require_tenant_context(request)
    if not gf_has_perm(request, "contracts.view"):
        raise PermissionDenied("Permission denied")
    can_write = gf_has_perm(request, "contracts.edit") or gf_has_perm(request, "contracts.create")

    if request.method == "GET":
        request.session.pop("finance_import_preview", None)
        return _render(request, can_write=can_write)
    if not can_write:
        raise PermissionDenied("Permission denied")

    action = str(request.POST.get("action") or "preview").strip()
    if action == "commit":
        preview = request.session.get("finance_import_preview") or {}
        selected = set(request.POST.getlist("selected"))
        if not preview or not selected:
            messages.warning(request, "저장할 행을 선택해 주세요.")
            return _render(request, can_write=can_write, preview=preview or None, import_type=preview.get("import_type", "invoice"))
        created, skipped = _commit_preview(alias, preview, selected, request)
        request.session.pop("finance_import_preview", None)
        messages.success(request, f"선택한 {created}건을 Finance에 저장했습니다.")
        if skipped:
            messages.info(request, f"선택하지 않은 항목 또는 저장 불가 항목 {skipped}건은 반영하지 않았습니다.")
        return redirect("tenant:finance_import")

    if action == "clear":
        request.session.pop("finance_import_preview", None)
        return redirect("tenant:finance_import")

    upload = request.FILES.get("file")
    import_type = str(request.POST.get("import_type") or "invoice").strip()
    invoice_direction = str(request.POST.get("invoice_direction") or "auto").strip()
    header_row_text = str(request.POST.get("header_row") or "").strip()
    if import_type not in {"invoice", "transaction"}:
        messages.error(request, "가져오기 유형을 확인하세요.")
        return _render(request, can_write=can_write)
    if not upload or not str(upload.name).lower().endswith(".xlsx"):
        messages.error(request, "XLSX 파일을 선택해 주세요.")
        return _render(request, can_write=can_write, import_type=import_type, invoice_direction=invoice_direction, header_row=header_row_text)

    try:
        manual_row = int(header_row_text) if header_row_text else None
        if manual_row is not None and manual_row < 1:
            raise ValueError("헤더 행은 1 이상의 숫자여야 합니다.")
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        detected_row, mapping, raw_headers = _find_header(ws, import_type, manual_row)
        if import_type == "invoice":
            rows = _invoice_preview(alias, ws, mapping, detected_row, invoice_direction)
        else:
            rows = _transaction_preview(alias, ws, mapping, detected_row)
        wb.close()
        matched = sorted(set(mapping).intersection(INVOICE_FIELDS if import_type == "invoice" else TRANSACTION_FIELDS))
        preview = {
            "import_type": import_type,
            "invoice_direction": invoice_direction,
            "header_row": detected_row,
            "matched_fields": matched,
            "file_name": str(upload.name),
            "rows": rows,
        }
        request.session["finance_import_preview"] = preview
        request.session.modified = True
        return _render(
            request,
            can_write=can_write,
            preview=preview,
            import_type=import_type,
            invoice_direction=invoice_direction,
            header_row=str(detected_row),
        )
    except Exception as exc:
        messages.error(request, f"Excel 미리보기에 실패했습니다: {exc}")
        return _render(request, can_write=can_write, import_type=import_type, invoice_direction=invoice_direction, header_row=header_row_text)
