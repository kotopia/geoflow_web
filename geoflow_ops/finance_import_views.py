from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import connections, transaction
from django.shortcuts import redirect, render
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook

from control.gf_authz.permissions import gf_has_perm
from .services.entity_access import require_tenant_context


HEADER_ALIASES = {
    "written_date": {"작성일", "작성일자"},
    "issued_date": {"발행일", "발행일자"},
    "invoice_type": {"구분", "매출매입구분", "계산서구분"},
    "partner": {"거래처", "거래처명", "업체명"},
    "contract": {"계약", "계약명", "계약번호"},
    "project": {"프로젝트", "프로젝트명", "프로젝트코드"},
    "supply_amount": {"공급가액", "공급가"},
    "vat_amount": {"부가세", "세액", "vat"},
    "total_amount": {"합계", "합계금액", "총액"},
    "approval_no": {"계산서번호", "승인번호", "계산서번호/승인번호", "전자세금계산서승인번호"},
    "status": {"상태"},
    "transaction_date": {"거래일자", "거래일", "일자"},
    "transaction_type": {"입출금구분", "입금출금구분", "구분"},
    "amount": {"금액", "거래금액"},
    "account": {"계좌", "계좌명", "계좌번호"},
    "description": {"적요", "내용"},
    "category": {"분류", "수입지출분류", "계정분류"},
    "evidence_type": {"증빙", "증빙유형"},
    "memo": {"비고", "메모"},
}


def _norm(value):
    return "".join(str(value or "").strip().lower().replace("·", "").replace("/", "").split())


def _headers(row):
    normalized = {_norm(v): idx for idx, v in enumerate(row) if v not in (None, "")}
    result = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = _norm(alias)
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def _cell(row, mapping, key):
    idx = mapping.get(key)
    return row[idx] if idx is not None and idx < len(row) else None


def _date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError("날짜 형식을 확인하세요")


def _money(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").replace("₩", "").strip())
    except (InvalidOperation, ValueError):
        raise ValueError("금액 형식을 확인하세요")


def _ref_code(alias, field_ref, value, required=False):
    text = str(value or "").strip()
    if not text:
        if required:
            raise ValueError("필수 참조값이 없습니다")
        return None
    with connections[alias].cursor() as cur:
        cur.execute("""
            SELECT v.code
              FROM ops.settings_nodes c JOIN ops.settings_nodes v ON v.parent_id=c.id
             WHERE c.field_ref=%s AND v.active=true
               AND (lower(v.code)=lower(%s) OR lower(v.name)=lower(%s))
             ORDER BY CASE WHEN lower(v.code)=lower(%s) THEN 0 ELSE 1 END, v.ord
             LIMIT 1
        """, [field_ref, text, text, text])
        row = cur.fetchone()
    if not row:
        raise ValueError(f"환경설정 참조값을 찾을 수 없습니다: {text}")
    return row[0]


def _lookup(alias, table, value, *, columns=("name", "code"), required=False):
    text = str(value or "").strip()
    if not text:
        if required:
            raise ValueError("필수 연결값이 없습니다")
        return None
    conditions = " OR ".join([f"lower(COALESCE({c}::text,''))=lower(%s)" for c in columns])
    params = [text] * len(columns)
    with connections[alias].cursor() as cur:
        cur.execute(f"SELECT id::text FROM {table} WHERE {conditions} LIMIT 2", params)
        rows = cur.fetchall()
    if not rows:
        raise ValueError(f"연결 대상을 찾을 수 없습니다: {text}")
    if len(rows) > 1:
        raise ValueError(f"동일한 연결 대상이 여러 건입니다: {text}")
    return rows[0][0]


def _project(alias, value, contract_id=None):
    text = str(value or "").strip()
    if not text:
        return None
    with connections[alias].cursor() as cur:
        if contract_id:
            cur.execute("""SELECT id::text FROM prj.projects WHERE contract_id=%s AND (lower(COALESCE(name,''))=lower(%s) OR lower(COALESCE(code,''))=lower(%s)) LIMIT 2""", [contract_id, text, text])
        else:
            cur.execute("""SELECT id::text FROM prj.projects WHERE lower(COALESCE(name,''))=lower(%s) OR lower(COALESCE(code,''))=lower(%s) LIMIT 2""", [text, text])
        rows = cur.fetchall()
    if not rows:
        raise ValueError(f"프로젝트를 찾을 수 없습니다: {text}")
    if len(rows) > 1:
        raise ValueError(f"프로젝트가 여러 건입니다: {text}")
    return rows[0][0]


def _account(alias, value):
    text = str(value or "").strip()
    if not text:
        return None
    with connections[alias].cursor() as cur:
        cur.execute("""SELECT id::text FROM fin.accounts WHERE active=true AND (lower(account_name)=lower(%s) OR replace(COALESCE(account_number,''),'-','')=replace(%s,'-','')) LIMIT 2""", [text, text])
        rows = cur.fetchall()
    if not rows:
        raise ValueError(f"계좌를 찾을 수 없습니다: {text}")
    if len(rows) > 1:
        raise ValueError(f"계좌가 여러 건입니다: {text}")
    return rows[0][0]


def _actor(request):
    user = request.user
    return str(getattr(user, "email", None) or getattr(user, "username", None) or getattr(user, "pk", ""))[:255]


def _import_invoices(alias, ws, mapping, request):
    created = 0
    errors = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in row):
            continue
        try:
            partner_id = _lookup(alias, "ctr.partners", _cell(row, mapping, "partner"), columns=("name",), required=True)
            contract_id = _lookup(alias, "ctr.contracts", _cell(row, mapping, "contract"), required=False)
            project_id = _project(alias, _cell(row, mapping, "project"), contract_id)
            invoice_type = _ref_code(alias, "finance.invoice_type", _cell(row, mapping, "invoice_type"), required=True)
            status_value = _cell(row, mapping, "status") or "발행"
            status = _ref_code(alias, "finance.invoice_status", status_value, required=True)
            supply = _money(_cell(row, mapping, "supply_amount"))
            vat = _money(_cell(row, mapping, "vat_amount"))
            total = _money(_cell(row, mapping, "total_amount")) if _cell(row, mapping, "total_amount") not in (None, "") else supply + vat
            approval_no = str(_cell(row, mapping, "approval_no") or "").strip() or None
            if approval_no:
                with connections[alias].cursor() as cur:
                    cur.execute("SELECT 1 FROM fin.tax_invoices WHERE approval_no=%s LIMIT 1", [approval_no])
                    if cur.fetchone():
                        continue
            with connections[alias].cursor() as cur:
                cur.execute("""INSERT INTO fin.tax_invoices(written_date,issued_date,invoice_type,partner_id,contract_id,project_id,supply_amount,vat_amount,total_amount,approval_no,status,memo,created_by)
                               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", [
                    _date(_cell(row,mapping,"written_date")), _date(_cell(row,mapping,"issued_date")), invoice_type, partner_id,
                    contract_id, project_id, supply, vat, total, approval_no, status, str(_cell(row,mapping,"memo") or "").strip() or None, _actor(request)
                ])
            created += 1
        except Exception as exc:
            errors.append(f"{row_no}행: {exc}")
    return created, errors


def _import_transactions(alias, ws, mapping, request):
    created = 0
    errors = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in row):
            continue
        try:
            contract_id = _lookup(alias, "ctr.contracts", _cell(row, mapping, "contract"), required=False)
            project_id = _project(alias, _cell(row, mapping, "project"), contract_id)
            partner_id = _lookup(alias, "ctr.partners", _cell(row, mapping, "partner"), columns=("name",), required=False)
            tx_type = _ref_code(alias, "finance.transaction_type", _cell(row, mapping, "transaction_type"), required=True)
            category = _ref_code(alias, "finance.transaction_category", _cell(row, mapping, "category"), required=False)
            evidence = _ref_code(alias, "finance.evidence_type", _cell(row, mapping, "evidence_type"), required=False)
            account_id = _account(alias, _cell(row, mapping, "account"))
            with connections[alias].cursor() as cur:
                cur.execute("""INSERT INTO fin.transactions(transaction_date,transaction_type,amount,partner_id,account_id,description,contract_id,project_id,category_code,evidence_type,memo,created_by)
                               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", [
                    _date(_cell(row,mapping,"transaction_date")), tx_type, _money(_cell(row,mapping,"amount")), partner_id, account_id,
                    str(_cell(row,mapping,"description") or "").strip() or None, contract_id, project_id, category, evidence,
                    str(_cell(row,mapping,"memo") or "").strip() or None, _actor(request)
                ])
            created += 1
        except Exception as exc:
            errors.append(f"{row_no}행: {exc}")
    return created, errors


@login_required
@require_http_methods(["GET", "POST"])
def finance_import(request):
    alias = require_tenant_context(request)
    if not gf_has_perm(request, "contracts.view"):
        raise PermissionDenied("Permission denied")
    can_write = gf_has_perm(request, "contracts.edit") or gf_has_perm(request, "contracts.create")
    if request.method == "GET":
        return render(request, "geoflow_ops/finance/finance_import.html", {})
    if not can_write:
        raise PermissionDenied("Permission denied")

    upload = request.FILES.get("file")
    import_type = str(request.POST.get("import_type") or "").strip()
    if not upload or import_type not in {"invoice", "transaction"}:
        messages.error(request, "가져오기 유형과 XLSX 파일을 확인하세요.")
        return redirect("tenant:finance_import")
    if not str(upload.name).lower().endswith(".xlsx"):
        messages.error(request, "XLSX 파일만 지원합니다.")
        return redirect("tenant:finance_import")

    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        mapping = _headers(first_row or [])
        required = {"invoice": {"invoice_type", "partner", "supply_amount"}, "transaction": {"transaction_date", "transaction_type", "amount"}}[import_type]
        missing = sorted(required - set(mapping))
        if missing:
            raise ValueError("필수 헤더가 없습니다: " + ", ".join(missing))
        with transaction.atomic(using=alias):
            if import_type == "invoice":
                created, errors = _import_invoices(alias, ws, mapping, request)
            else:
                created, errors = _import_transactions(alias, ws, mapping, request)
        wb.close()
    except Exception as exc:
        messages.error(request, f"Excel 가져오기에 실패했습니다: {exc}")
        return redirect("tenant:finance_import")

    messages.success(request, f"{created}건을 가져왔습니다.")
    if errors:
        request.session["finance_import_errors"] = errors[:100]
        messages.warning(request, f"{len(errors)}건은 가져오지 못했습니다. 아래 오류 목록을 확인하세요.")
    return redirect("tenant:finance_import")
